from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import re
from sklearn.ensemble import VotingClassifier

import warnings
warnings.filterwarnings("ignore")
plt.style.use('ggplot')
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics import accuracy_score, confusion_matrix, classification_report
from sklearn.metrics import accuracy_score
from sklearn.metrics import f1_score
from sklearn.tree import DecisionTreeClassifier

# Create your views here.
from Remote_User.models import ClientRegister_Model,Fraudsters_and_Fraudulent_details,Fraudsters_and_Fraudulent_prediction,detection_ratio_model,detection_accuracy_model

def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            enter = ClientRegister_Model.objects.get(username=username,password=password)
            request.session["userid"] = enter.id

            return redirect('Add_DataSet_Details')
        except:
            pass

    return render(request,'RUser/login.html')

def Add_DataSet_Details(request):
    if "GET" == request.method:
        return render(request, 'RUser/Add_DataSet_Details.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)
        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)
        # reading a cell
        print(worksheet["A1"].value)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)
            Fraudsters_and_Fraudulent_details.objects.all().delete()
    for r in range(1, active_sheet.max_row+1):
        Fraudsters_and_Fraudulent_details.objects.create(
        step= active_sheet.cell(r, 1).value,
        type= active_sheet.cell(r, 2).value,
        amount= active_sheet.cell(r, 3).value,
        nameOrig= active_sheet.cell(r, 4).value,
        oldbalanceOrg= active_sheet.cell(r, 5).value,
        newbalanceOrig= active_sheet.cell(r, 6).value,
        nameDest= active_sheet.cell(r, 7).value,
        oldbalanceDest= active_sheet.cell(r, 8).value,
        newbalanceDest= active_sheet.cell(r, 9).value,
        isFraud= active_sheet.cell(r, 10).value
        )

    return render(request, 'RUser/Add_DataSet_Details.html', {"excel_data": excel_data})


def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:
        return render(request,'RUser/Register1.html')

def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})


def Predict_fraudsters_and_fraudulent(request):
        expense = 0
        kg_price=0
        if request.method == "POST":

            ptype = request.POST.get('ptype')
            amount = request.POST.get('amount')
            tfrom = request.POST.get('tfrom')
            obal = request.POST.get('obal')
            nbal=request.POST.get('nbal')
            Tto = request.POST.get('Tto')
            ROb = request.POST.get('ROb')
            RNb = request.POST.get('RNb')

            df = pd.read_csv('DataSets.csv')
            df
            df.columns
            df.rename(columns={'isFraud': 'label', 'nameDest': 'CustomerId'}, inplace=True)

            def apply_results(label):
                if (label == 0):
                    return 0  # non fraudulent transaction
                else:
                    return 1  # fraudulent transaction

            df['results'] = df['label'].apply(apply_results)
            df.drop(['label'], axis=1, inplace=True)
            results = df['results'].value_counts()
            df.drop(['step', 'type', 'amount', 'nameOrig', 'oldbalanceOrg', 'newbalanceOrig', 'oldbalanceDest',
                     'newbalanceDest', 'isFlaggedFraud'], axis=1, inplace=True)

            cv = CountVectorizer()
            X = df['CustomerId']
            y = df['results']

            print("Customer Id")
            print(X)
            print("Results")
            print(y)

            X = cv.fit_transform(X)

            models = []
            from sklearn.model_selection import train_test_split
            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.20)
            X_train.shape, X_test.shape, y_train.shape

            print("Naive Bayes")

            from sklearn.naive_bayes import MultinomialNB
            NB = MultinomialNB()
           # SVM Model
            print("SVM")
            from sklearn import svm
            lin_clf = svm.LinearSVC()


            print("Logistic Regression")

            from sklearn.linear_model import LogisticRegression
            reg = LogisticRegression(random_state=0, solver='lbfgs').fit(X_train, y_train)

            print("Decision Tree Classifier")
            dtc = DecisionTreeClassifier()

            obj=Fraudsters_and_Fraudulent_details.objects.get(type=ptype,amount=amount,nameOrig=tfrom,oldbalanceOrg=obal,newbalanceOrig=nbal,nameDest=Tto,oldbalanceDest=ROb,newbalanceDest=RNb)
            ttype1 =int(obj.isFraud)
            ttype=apply_results(ttype1)
            step=1;
            print(obj.isFraud)

            if int(ttype)==0:
               val="Non Fraud"
            elif int(ttype) == 1:
                val = "Fraud"

            Fraudsters_and_Fraudulent_prediction.objects.create(step=step,type=ptype,amount=amount,nameOrig=tfrom,oldbalanceOrg=obal,newbalanceOrig=nbal,nameDest=Tto,oldbalanceDest=ROb,newbalanceDest=RNb,prediction=val)


            return render(request, 'RUser/Predict_fraudsters_and_fraudulent.html',{'objs':val})
        return render(request, 'RUser/Predict_fraudsters_and_fraudulent.html')


def ratings(request,pk):
    vott1, vott, neg = 0, 0, 0
    objs = Fraudsters_and_Fraudulent_details.objects.get(id=pk)
    unid = objs.id
    vot_count = Fraudsters_and_Fraudulent_details.objects.all().filter(id=unid)
    for t in vot_count:
        vott = t.ratings
        vott1 = vott + 1
        obj = get_object_or_404(Fraudsters_and_Fraudulent_details, id=unid)
        obj.ratings = vott1
        obj.save(update_fields=["ratings"])
        return redirect('Add_DataSet_Details')

    return render(request,'RUser/ratings.html',{'objs':vott1})



